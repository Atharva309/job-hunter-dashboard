import os
import sys
import time
import json
import base64
import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import PatternFill, Font
from anthropic import Anthropic

EXCEL_FILE = "job_tracker.xlsx"
PDF_FILE = "portfolio.pdf"
MODEL_NAME = "claude-sonnet-4-20250514"

# Will be initialized inside process_jobs
client = None

SYSTEM_PROMPT = """
You are a job search assistant helping an F1 international student named Atharva Patil find entry-level jobs in the USA.

Candidate profile summary:
- MS Data Science @ CU Boulder (graduating May 2026), GPA 3.9
- Skills: Python, R, SQL, ML/DL, NLP, Computer Vision, ETL, FastAPI, AWS, Power BI, Tableau
- Target roles: Data Science, ML Engineer, Data Engineer, AI Engineer, Analytics, Software Engineer (backend/data), Business Analyst
- NEEDS visa sponsorship (F1 OPT → H1B). Prioritize companies known to sponsor.

You will receive scraped text from a company's career page.

Find up to 3 best-matching job postings. Apply these strict filters:
✓ USA only (remote-USA is fine, international = skip)
✓ Entry-level / new grad only (skip: senior, staff, lead, manager, director, principal)
✓ Must match Atharva's target roles above
✓ Posted within last 60 days if date is visible (skip older)

Return ONLY a raw JSON array, no markdown, no explanation:
[
  {
    "job_title": "...",
    "apply_link": "...",
    "location": "...",
    "sponsorship": "Yes | No | Not Mentioned",
    "entry_level": "Yes | No",
    "date_posted": "YYYY-MM-DD or Not Listed",
    "match_score": 8,
    "notes": "one sentence on why this fits Atharva"
  }
]

If zero jobs match all filters, return exactly: []
"""

def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Jobs"
        
        headers = [
            "Company Name", "Career Page URL", "Matched Job Title", 
            "Job Apply Link", "Location", "Sponsorship Mentioned?", 
            "Entry Level?", "Date Posted", "Match Score 1-10", 
            "Status", "Notes"
        ]
        ws.append(headers)
        
        dark_blue_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        light_blue_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        light_green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        white_font = Font(color="FFFFFF", bold=True)
        
        # Apply header formatting
        for col_idx, cell in enumerate(ws[1], 1):
            cell.fill = dark_blue_fill
            cell.font = white_font
            
        # Freeze top row
        ws.freeze_panes = "A2"
        wb.save(EXCEL_FILE)
        print(f"Created {EXCEL_FILE}. Please fill in columns A and B, then run this script again.")
        sys.exit(0)

def scrape_url(url):
    headers = {
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.5",
        "DNT": "1",
        "Connection": "keep-alive",
        "Upgrade-Insecure-Requests": "1"
    }
    try:
        response = requests.get(url, headers=headers, timeout=10)
        response.raise_for_status()
        soup = BeautifulSoup(response.text, "html.parser")
        
        # Remove script, style, nav, footer
        for el in soup(["script", "style", "nav", "footer", "header", "noscript"]):
            el.decompose()
        # Extract text with markdown-style links so Claude can see them
        text_parts = []
        for element in soup.descendants:
            if isinstance(element, str):
                cleaned = element.strip()
                if cleaned:
                    text_parts.append(cleaned)
            elif element.name == "a" and element.get("href"):
                # Append the href next to the link text
                text_parts.append(f"({element.get('href')})")
                
        text = " ".join(text_parts)
        return text[:6000]
    except Exception as e:
        return None

def analyze_with_claude(scraped_text, pdf_b64):
    global client
    try:
        if not client:
            client = Anthropic()
        
        message = client.messages.create(
            model=MODEL_NAME,
            max_tokens=1000,
            system=SYSTEM_PROMPT,
            messages=[
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "document",
                            "source": {
                                "type": "base64",
                                "media_type": "application/pdf",
                                "data": pdf_b64
                            }
                        },
                        {
                            "type": "text",
                            "text": f"Scraped career page text for analysis:\n\n{scraped_text}"
                        }
                    ]
                }
            ]
        )
        # Parse output safely
        content = message.content[0].text.strip()
        start_idx = content.find('[')
        end_idx = content.rfind(']')
        if start_idx != -1 and end_idx != -1:
            content = content[start_idx:end_idx+1]
        return json.loads(content)
    except Exception as e:
        print(f"Claude API Error: {e}")
        return None

def process_jobs():
    global client
    
    # Ensure API key is set before running
    if not os.environ.get("ANTHROPIC_API_KEY"):
        msg = "Error: ANTHROPIC_API_KEY environment variable not set."
        print(msg)
        return {"status": "error", "message": msg}
        
    init_excel()
    
    if not os.path.exists(PDF_FILE):
        msg = f"Error: {PDF_FILE} not found. Please place your portfolio PDF in the directory."
        print(msg)
        return {"status": "error", "message": msg}
        
    with open(PDF_FILE, "rb") as f:
        pdf_b64 = base64.b64encode(f.read()).decode("utf-8")
        
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    
    # We will process rows backwards so that inserting/appending rows doesn't mess up the loop index,
    # or we can collect tasks first. Let's collect rows to process.
    rows_to_process = []
    for r in range(2, ws.max_row + 1):
        company = ws.cell(row=r, column=1).value
        url = ws.cell(row=r, column=2).value
        matched = ws.cell(row=r, column=3).value
        status = ws.cell(row=r, column=10).value
        # Skip applied jobs — they're done
        if status == "Applied":
            continue
        # If 'matched' is empty or contains an error from a previous run, retry it
        if company and url and (not matched or "Error" in str(matched) or "Failed" in str(matched)):
            rows_to_process.append((r, company, url))
            
    if not rows_to_process:
        msg = "No new companies to process."
        print(msg)
        return {"status": "success", "message": msg}

    # Track row shifts if we insert rows
    offset = 0

    for original_row, company, url in rows_to_process:
        curr_row = original_row + offset
        
        # 1. Scrape
        text = scrape_url(url)
        # Try fallbacks if first scrape fails or is too short
        if not text or len(text) < 200:
            for suffix in ["/jobs", "/careers", "/openings"]:
                base = url.rstrip('/')
                fallback_text = scrape_url(base + suffix)
                if fallback_text and len(fallback_text) > 200:
                    text = fallback_text
                    break
                    
        if not text:
            ws.cell(row=curr_row, column=3).value = "Failed to scrape"
            ws.cell(row=curr_row, column=11).value = "Failed to scrape page."
            wb.save(EXCEL_FILE)
            print(f"✗ {company} → Failed to scrape target page")
            continue
            
        # 2. Claude API
        results = analyze_with_claude(text, pdf_b64)
        
        if results is None:
            ws.cell(row=curr_row, column=3).value = "Error parsing from Claude"
            wb.save(EXCEL_FILE)
            continue
            
        # 3. Filter out already-applied jobs
        applied_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "applied.json")
        applied_titles = set()
        if os.path.exists(applied_file):
            with open(applied_file, "r") as af:
                for entry in json.load(af):
                    if entry.get("company") == company:
                        applied_titles.add(entry.get("title", "").lower().strip())
        
        results = [j for j in results if j.get("job_title", "").lower().strip() not in applied_titles]
        
        # 4. Write results
        if len(results) == 0:
            ws.cell(row=curr_row, column=3).value = "No matches found"
            print(f"✗ {company} → No matches")
        else:
            print(f"✓ {company} → {len(results)} jobs found")
            for i, job in enumerate(results):
                target_row = curr_row + i
                if i > 0:
                    ws.insert_rows(target_row)
                    offset += 1
                    # Copy company and URL to the inserted row
                    ws.cell(row=target_row, column=1).value = company
                    ws.cell(row=target_row, column=2).value = url
                
                # Apply light blue formatting to input cols for inserted rows if needed
                for col in [1, 2]:
                    ws.cell(row=target_row, column=col).fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
                
                ws.cell(row=target_row, column=3).value = job.get("job_title", "")
                ws.cell(row=target_row, column=4).value = job.get("apply_link", "")
                ws.cell(row=target_row, column=5).value = job.get("location", "")
                ws.cell(row=target_row, column=6).value = job.get("sponsorship", "")
                ws.cell(row=target_row, column=7).value = job.get("entry_level", "")
                ws.cell(row=target_row, column=8).value = job.get("date_posted", "")
                ws.cell(row=target_row, column=9).value = job.get("match_score", "")
                ws.cell(row=target_row, column=11).value = job.get("notes", "")
                
                for col in range(3, 12):
                    if col != 10: # Status col (user fills)
                        ws.cell(row=target_row, column=col).fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                        
            wb.save(EXCEL_FILE)
            
        time.sleep(2) # polite delay

    return {"status": "success", "message": "Scraping complete!"}

def process_single_company(company_name):
    """Scan a single company by name. Reuses the same logic as process_jobs but for one row."""
    global client
    
    if not os.environ.get("ANTHROPIC_API_KEY"):
        return {"status": "error", "message": "ANTHROPIC_API_KEY not set."}
    
    init_excel()
    
    if not os.path.exists(PDF_FILE):
        return {"status": "error", "message": f"{PDF_FILE} not found."}
    
    with open(PDF_FILE, "rb") as f:
        pdf_b64 = base64.b64encode(f.read()).decode("utf-8")
    
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    
    # Find the FIRST non-applied row for this company, or get URL from any row
    target_row = None
    company_url = None
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == company_name:
            if not company_url:
                company_url = ws.cell(row=r, column=2).value
            if ws.cell(row=r, column=10).value != "Applied":
                target_row = r
                break
    
    if not company_url:
        return {"status": "error", "message": f"Company '{company_name}' not found in tracker."}
    
    # If all rows are applied, insert a fresh row for new results
    if not target_row:
        new_row = ws.max_row + 1
        ws.cell(row=new_row, column=1).value = company_name
        ws.cell(row=new_row, column=2).value = company_url
        target_row = new_row
        wb.save(EXCEL_FILE)
    
    url = company_url
    
    # Clear previous results on this row (but preserve the status column)
    for col in range(3, 12):
        if col != 10:
            ws.cell(row=target_row, column=col).value = None
    wb.save(EXCEL_FILE)
    
    # Scrape
    text = scrape_url(url)
    if not text or len(text) < 200:
        for suffix in ["/jobs", "/careers", "/openings"]:
            base = url.rstrip('/')
            fallback_text = scrape_url(base + suffix)
            if fallback_text and len(fallback_text) > 200:
                text = fallback_text
                break
    
    if not text:
        ws.cell(row=target_row, column=3).value = "Failed to scrape"
        ws.cell(row=target_row, column=11).value = "Failed to scrape page."
        wb.save(EXCEL_FILE)
        return {"status": "error", "message": f"Failed to scrape {company_name}"}
    
    # Claude API
    results = analyze_with_claude(text, pdf_b64)
    
    if results is None:
        ws.cell(row=target_row, column=3).value = "Error parsing from Claude"
        wb.save(EXCEL_FILE)
        return {"status": "error", "message": f"Claude failed to parse results for {company_name}"}
    
    if len(results) == 0:
        ws.cell(row=target_row, column=3).value = "No matches found"
        wb.save(EXCEL_FILE)
        return {"status": "success", "message": f"{company_name}: No matching jobs found."}
    
    # Filter out jobs the user has already applied to
    applied_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "applied.json")
    applied_titles = set()
    if os.path.exists(applied_file):
        with open(applied_file, "r") as af:
            for entry in json.load(af):
                if entry.get("company") == company_name:
                    applied_titles.add(entry.get("title", "").lower().strip())
    
    results = [j for j in results if j.get("job_title", "").lower().strip() not in applied_titles]
    
    if len(results) == 0:
        # Delete the temp row so it doesn't show as junk in the UI
        ws.delete_rows(target_row)
        wb.save(EXCEL_FILE)
        return {"status": "success", "message": f"{company_name}: All found jobs were already applied to."}
    
    # Write results
    for i, job in enumerate(results):
        row = target_row + i
        if i > 0:
            ws.insert_rows(row)
            ws.cell(row=row, column=1).value = company_name
            ws.cell(row=row, column=2).value = url
        
        for col in [1, 2]:
            ws.cell(row=row, column=col).fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        
        ws.cell(row=row, column=3).value = job.get("job_title", "")
        ws.cell(row=row, column=4).value = job.get("apply_link", "")
        ws.cell(row=row, column=5).value = job.get("location", "")
        ws.cell(row=row, column=6).value = job.get("sponsorship", "")
        ws.cell(row=row, column=7).value = job.get("entry_level", "")
        ws.cell(row=row, column=8).value = job.get("date_posted", "")
        ws.cell(row=row, column=9).value = job.get("match_score", "")
        ws.cell(row=row, column=11).value = job.get("notes", "")
        
        for col in range(3, 12):
            if col != 10:
                ws.cell(row=row, column=col).fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    
    wb.save(EXCEL_FILE)
    return {"status": "success", "message": f"{company_name}: {len(results)} job(s) found!"}

if __name__ == "__main__":
    process_jobs()
