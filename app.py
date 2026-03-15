from fastapi import FastAPI, HTTPException
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse
from pydantic import BaseModel
import openpyxl
import os
import json
import pathlib
import traceback

# Import the core agent logic
from job_agent import process_jobs, process_single_company, EXCEL_FILE, init_excel

app = FastAPI(title="Job Hunter Dashboard")

BASE_DIR = pathlib.Path(__file__).parent.resolve()
APPLIED_FILE = str(BASE_DIR / "applied.json")

# Ensure excel exists on startup
if not os.path.exists(EXCEL_FILE):
    init_excel()

app.mount("/static", StaticFiles(directory=str(BASE_DIR / "static")), name="static")

class CompanyPayload(BaseModel):
    name: str
    url: str

class AppliedPayload(BaseModel):
    company: str
    title: str
    apply_link: str = ""
    location: str = ""

# --- Helpers ---
def load_applied():
    if os.path.exists(APPLIED_FILE):
        with open(APPLIED_FILE, "r") as f:
            return json.load(f)
    return []

def save_applied(data):
    with open(APPLIED_FILE, "w") as f:
        json.dump(data, f, indent=2)

# --- Routes ---
@app.get("/")
def serve_index():
    return FileResponse(str(BASE_DIR / "static" / "index.html"))

@app.get("/api/jobs")
def get_jobs():
    if not os.path.exists(EXCEL_FILE):
        return []
    
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    jobs = []
    
    for r in range(2, ws.max_row + 1):
        matched = ws.cell(row=r, column=3).value
        status = ws.cell(row=r, column=10).value
        # Skip applied jobs and non-matches
        if status == "Applied":
            continue
        if matched and "No matches" not in str(matched) and "Error" not in str(matched) and "Failed" not in str(matched):
            jobs.append({
                "company": ws.cell(row=r, column=1).value,
                "url": ws.cell(row=r, column=2).value,
                "title": matched,
                "apply_link": ws.cell(row=r, column=4).value,
                "location": ws.cell(row=r, column=5).value,
                "sponsorship": ws.cell(row=r, column=6).value,
                "entry_level": ws.cell(row=r, column=7).value,
                "date_posted": ws.cell(row=r, column=8).value,
                "score": ws.cell(row=r, column=9).value,
                "notes": ws.cell(row=r, column=11).value
            })
    return jobs

@app.get("/api/companies")
def get_companies():
    if not os.path.exists(EXCEL_FILE):
        return []
        
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    companies = []
    
    for r in range(2, ws.max_row + 1):
        name = ws.cell(row=r, column=1).value
        url = ws.cell(row=r, column=2).value
        
        if name and not any(c['name'] == name for c in companies):
            matched = ws.cell(row=r, column=3).value
            status = "Pending"
            if matched == "No matches found":
                status = "No Matches"
            elif matched and ("Error" in str(matched) or "Failed" in str(matched)):
                status = "Error"
            elif matched:
                status = "Found Jobs"
                
            companies.append({
                "name": name,
                "url": url,
                "status": status
            })
            
    return companies

@app.post("/api/add-company")
def add_company(payload: CompanyPayload):
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        ws.append([payload.name, payload.url] + [""] * 9)
        wb.save(EXCEL_FILE)
        return {"status": "success", "message": f"Added {payload.name}"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/companies/{company_name}")
def delete_company(company_name: str):
    """Remove all rows for a company from the Excel tracker."""
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        rows_to_delete = []
        for r in range(2, ws.max_row + 1):
            if ws.cell(row=r, column=1).value == company_name:
                rows_to_delete.append(r)
        
        if not rows_to_delete:
            raise HTTPException(status_code=404, detail=f"Company '{company_name}' not found")
        
        # Delete from bottom to top so row indices don't shift
        for r in sorted(rows_to_delete, reverse=True):
            ws.delete_rows(r)
        
        wb.save(EXCEL_FILE)
        return {"status": "success", "message": f"Removed {company_name}"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/scan")
def trigger_scan():
    """Run the job scan for ALL pending companies."""
    try:
        result = process_jobs()
        if result is None:
            return {"status": "success", "message": "Scan finished (no status returned)."}
        if result.get("status") == "error":
            return {"status": "error", "message": result.get("message", "Unknown error")}
        return {"status": "success", "message": result.get("message", "Scan complete")}
    except Exception as e:
        tb = traceback.format_exc()
        print(f"SCAN ERROR: {tb}")
        return {"status": "error", "message": str(e)}

@app.post("/api/scan/{company_name}")
def scan_single(company_name: str):
    """Run the job scan for a SINGLE company."""
    try:
        result = process_single_company(company_name)
        if result is None:
            return {"status": "success", "message": "Scan finished (no status returned)."}
        if result.get("status") == "error":
            return {"status": "error", "message": result.get("message", "Unknown error")}
        return {"status": "success", "message": result.get("message", "Scan complete")}
    except Exception as e:
        tb = traceback.format_exc()
        print(f"SCAN ERROR: {tb}")
        return {"status": "error", "message": str(e)}

# --- Applied Jobs ---
@app.get("/api/applied")
def get_applied():
    data = load_applied()
    return {"count": len(data), "jobs": data}

@app.post("/api/applied")
def mark_applied(payload: AppliedPayload):
    # 1. Save to applied.json
    data = load_applied()
    entry = {
        "company": payload.company,
        "title": payload.title,
        "apply_link": payload.apply_link,
        "location": payload.location,
        "applied_at": __import__('datetime').datetime.now().isoformat()
    }
    data.append(entry)
    save_applied(data)
    
    # 2. Mark as "Applied" in Excel column 10 so it won't show in Jobs or get re-scanned
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        for r in range(2, ws.max_row + 1):
            if ws.cell(row=r, column=1).value == payload.company and ws.cell(row=r, column=3).value == payload.title:
                ws.cell(row=r, column=10).value = "Applied"
                break
        wb.save(EXCEL_FILE)
    except Exception:
        pass  # Non-critical, the job is still tracked in applied.json
    
    return {"status": "success", "count": len(data), "message": f"Marked as applied: {payload.title} @ {payload.company}"}

@app.delete("/api/applied/{index}")
def remove_applied(index: int):
    data = load_applied()
    if index < 0 or index >= len(data):
        raise HTTPException(status_code=404, detail="Invalid index")
    removed = data.pop(index)
    save_applied(data)
    return {"status": "success", "count": len(data), "message": f"Removed {removed['title']}"}

if __name__ == "__main__":
    import uvicorn
    port = int(os.environ.get("PORT", 5050))
    uvicorn.run(app, host="0.0.0.0", port=port)
